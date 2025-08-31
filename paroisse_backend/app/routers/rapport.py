from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from typing import List, Optional
from fastapi.responses import FileResponse
from reportlab.pdfgen import canvas
from openpyxl import Workbook
import tempfile
from datetime import datetime

from app.schemas.rapport import RapportCreate, RapportUpdate, RapportOut
from app.crud import rapport as crud_rapport
from app.database import get_db
from app.utils.security import get_current_user

from app.permissions.rapport import (
    ALLOWED_ROLES_FINANCIER,
    ALLOWED_ROLES_ADMINISTRATIF,
    ALLOWED_ROLES_AUDIT,
    ALLOWED_ROLES_MATERIEL
)

router = APIRouter(prefix="/rapport", tags=["Rapport"])

def check_role(user, allowed_roles):
    if user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="Accès refusé: rôle non autorisé")

# --- ROUTES GÉNÉRALES ---

@router.get("/", response_model=List[RapportOut])
def list_rapports(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    include_deleted: bool = False
):
    check_role(current_user, ALLOWED_ROLES_FINANCIER | ALLOWED_ROLES_ADMINISTRATIF | ALLOWED_ROLES_AUDIT | ALLOWED_ROLES_MATERIEL)
    return crud_rapport.get_rapports(db, include_deleted=include_deleted)

@router.get("/{rapport_id}", response_model=RapportOut)
def get_rapport(
    rapport_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, ALLOWED_ROLES_FINANCIER | ALLOWED_ROLES_ADMINISTRATIF | ALLOWED_ROLES_AUDIT | ALLOWED_ROLES_MATERIEL)
    rapport = crud_rapport.get_rapport(db, rapport_id)
    if not rapport:
        raise HTTPException(status_code=404, detail="Rapport non trouvé")
    return rapport

@router.put("/{rapport_id}", response_model=RapportOut)
def update_rapport(
    rapport_id: int,
    rapport: RapportUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, ALLOWED_ROLES_FINANCIER | ALLOWED_ROLES_ADMINISTRATIF | ALLOWED_ROLES_AUDIT | ALLOWED_ROLES_MATERIEL)
    rapport_modifie = crud_rapport.update_rapport(db, rapport_id, rapport)
    if not rapport_modifie:
        raise HTTPException(status_code=404, detail="Rapport non trouvé")
    return rapport_modifie

@router.delete("/{rapport_id}")
def soft_delete_rapport(
    rapport_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, ALLOWED_ROLES_FINANCIER | ALLOWED_ROLES_ADMINISTRATIF | ALLOWED_ROLES_AUDIT | ALLOWED_ROLES_MATERIEL)
    rapport = crud_rapport.soft_delete_rapport(db, rapport_id)
    if not rapport:
        raise HTTPException(status_code=404, detail="Rapport non trouvé")
    return {"message": "Rapport mis dans la corbeille"}

@router.put("/restore/{rapport_id}")
def restore_rapport(
    rapport_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, ALLOWED_ROLES_FINANCIER | ALLOWED_ROLES_ADMINISTRATIF | ALLOWED_ROLES_AUDIT | ALLOWED_ROLES_MATERIEL)
    rapport = crud_rapport.restore_rapport(db, rapport_id)
    if not rapport:
        raise HTTPException(status_code=404, detail="Rapport non trouvé ou non supprimé")
    return {"message": "Rapport restauré"}

# --- CRÉATION DES RAPPORTS PAR TYPE ---

@router.post("/financier", response_model=RapportOut, status_code=status.HTTP_201_CREATED)
def create_rapport_financier(
    rapport: RapportCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, ALLOWED_ROLES_FINANCIER)
    if rapport.type != "financier":
        raise HTTPException(status_code=403, detail="Seuls les rapports financiers sont autorisés ici.")
    return crud_rapport.create_rapport_financier(db, rapport, current_user.utilisateur_id)

@router.post("/administratif", response_model=RapportOut, status_code=status.HTTP_201_CREATED)
def create_rapport_administratif(
    rapport: RapportCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, ALLOWED_ROLES_ADMINISTRATIF)
    if rapport.type != "administratif":
        raise HTTPException(status_code=403, detail="Seuls les rapports administratifs sont autorisés ici.")
    return crud_rapport.create_rapport_administratif(db, rapport, current_user.utilisateur_id)

@router.post("/materiel", response_model=RapportOut, status_code=status.HTTP_201_CREATED)
def create_rapport_materiel(
    rapport: RapportCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, ALLOWED_ROLES_MATERIEL)
    if rapport.type != "materiel":
        raise HTTPException(status_code=403, detail="Seuls les rapports matériels sont autorisés ici.")
    return crud_rapport.create_rapport_materiel(db, rapport, current_user.utilisateur_id)

@router.post("/audit", response_model=RapportOut, status_code=status.HTTP_201_CREATED)
def create_rapport_audit(
    rapport: RapportCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, ALLOWED_ROLES_AUDIT)
    if rapport.type != "audit":
        raise HTTPException(status_code=403, detail="Seuls les rapports d'audit sont autorisés ici.")
    return crud_rapport.create_rapport_audit(db, rapport, current_user.utilisateur_id)


# --- RAPPORT FINANCIER AUTOMATIQUE ---

@router.get("/budget-annuel/{annee}")
def rapport_budget_annuel(
    annee: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, ALLOWED_ROLES_FINANCIER)
    return crud_rapport.generer_rapport_financier_annuel(db, annee, current_user.utilisateur_id)

@router.get("/export-pdf/{annee}")
def export_pdf_financier(
    annee: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, ALLOWED_ROLES_FINANCIER)
    data = crud_rapport.generer_rapport_financier_annuel(db, annee, current_user.utilisateur_id)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        c = canvas.Canvas(tmp.name)
        c.setFont("Helvetica", 14)
        c.drawString(100, 800, f"Rapport Budgétaire - Année {annee}")
        y = 780
        for section, content in data.items():
            if isinstance(content, dict):
                c.drawString(100, y, f"{section.capitalize()}:")
                y -= 20
                for key, val in content.items():
                    if isinstance(val, dict):
                        c.drawString(120, y, f"{key.capitalize()}:")
                        y -= 15
                        for k2, v2 in val.items():
                            c.drawString(140, y, f"{k2}: {v2}")
                            y -= 15
                    else:
                        c.drawString(120, y, f"{key}: {val}")
                        y -= 15
            else:
                c.drawString(100, y, f"{section}: {content}")
                y -= 20
        c.save()
        tmp_path = tmp.name
    return FileResponse(tmp_path, filename=f"rapport_budget_{annee}.pdf", media_type="application/pdf")

@router.get("/export-excel/{annee}")
def export_excel_financier(
    annee: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, ALLOWED_ROLES_FINANCIER)
    data = crud_rapport.generer_rapport_financier_annuel(db, annee, current_user.utilisateur_id)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Budget {annee}"
    ws.append(["Catégorie", "Montant (FCFA)"])
    ws.append(["Recettes totales", data["recettes"]["total"]])
    for k, v in data["recettes"]["details"].items():
        ws.append([f"  {k.capitalize()}", v])
    ws.append(["Dépenses totales", data["depenses"]["total"]])
    for k, v in data["depenses"]["details"].items():
        ws.append([f"  {k.capitalize()}", v])
    ws.append(["Budget prévisionnel", data["budget"]["previsionnel"]])
    ws.append(["Budget réel", data["budget"]["reel"]])
    ws.append(["Écart", data["budget"]["ecart"]])
    ws.append(["Solde", data["solde"]])
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    return FileResponse(tmp_path, filename=f"rapport_budget_{annee}.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@router.get("/administratif/export-pdf")
def export_pdf_rapport_administratif(
    date_debut: Optional[datetime] = None,
    date_fin: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, ALLOWED_ROLES_ADMINISTRATIF)
    data = crud_rapport.generer_rapport_administratif(db, date_debut, date_fin)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        c = canvas.Canvas(tmp.name)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(100, 800, "Rapport Administratif")
        y = 780

        c.setFont("Helvetica", 12)
        c.drawString(100, y, f"Total rapports: {data['total_rapports']}")
        y -= 30

        for rapport in data["rapports"]:
            c.setFont("Helvetica-Bold", 12)
            c.drawString(100, y, f"Titre: {rapport['titre']}")
            y -= 15
            c.setFont("Helvetica-Oblique", 10)
            c.drawString(100, y, f"Date: {rapport['date_rapport']}  |  Auteur: {rapport['auteur']}")
            y -= 15
            c.setFont("Helvetica", 10)
            for line in rapport["contenu"].split("\n"):
                c.drawString(110, y, line)
                y -= 12
                if y < 50:
                    c.showPage()
                    y = 800
            y -= 20
            if y < 50:
                c.showPage()
                y = 800

        c.save()
        tmp_path = tmp.name

    return FileResponse(tmp_path, filename="rapport_administratif.pdf", media_type="application/pdf")


@router.get("/administratif/export-excel")
def export_excel_rapport_administratif(
    date_debut: Optional[datetime] = None,
    date_fin: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, ALLOWED_ROLES_ADMINISTRATIF)
    data = crud_rapport.generer_rapport_administratif(db, date_debut, date_fin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Administratif"
    ws.append(["Titre", "Date Rapport", "Auteur", "Contenu"])

    for rapport in data["rapports"]:
        ws.append([rapport["titre"], rapport["date_rapport"], rapport["auteur"], rapport["contenu"]])

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    return FileResponse(
        tmp_path,
        filename="rapport_administratif.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )




@router.get("/audit/compile/export-pdf")
def export_pdf_rapport_audit_compilé(
    date_debut: Optional[datetime] = None,
    date_fin: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, {"Administrateur", "Inspecteur"})
    if not date_debut or not date_fin:
        raise HTTPException(status_code=400, detail="date_debut et date_fin sont requis.")

    data = crud_rapport.generer_rapport_audit_compile(db, date_debut, date_fin)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        c = canvas.Canvas(tmp.name)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(100, 800, "Rapport d'Audit Combiné")
        y = 780

        # Période
        c.setFont("Helvetica", 12)
        c.drawString(100, y, f"Période: {data['periode']['date_debut']} à {data['periode']['date_fin']}")
        y -= 30

        # ADMINISTRATIF
        c.setFont("Helvetica-Bold", 14)
        c.drawString(100, y, f"Rapport Administratif - Total: {data['administratif']['total_rapports']}")
        y -= 20
        for rapport in data['administratif']['rapports']:
            c.setFont("Helvetica-Bold", 12)
            c.drawString(100, y, f"Titre: {rapport['titre']}")
            y -= 15
            c.setFont("Helvetica-Oblique", 10)
            c.drawString(100, y, f"Date: {rapport['date_rapport']} | Auteur: {rapport['auteur']}")
            y -= 15
            c.setFont("Helvetica", 10)
            for line in rapport['contenu'].split('\n'):
                c.drawString(110, y, line)
                y -= 12
                if y < 50:
                    c.showPage()
                    y = 800
            y -= 20

        # FINANCIER
        c.setFont("Helvetica-Bold", 14)
        c.drawString(100, y, f"Rapport Financier - Année: {data['periode']['annee_financiere']}")
        y -= 20
        recettes = data['financier'].get('recettes', {})
        depenses = data['financier'].get('depenses', {})
        budget = data['financier'].get('budget', {})
        solde = data['financier'].get('solde', 0)

        c.setFont("Helvetica", 12)
        c.drawString(100, y, f"Recettes totales: {recettes.get('total', 0)} FCFA")
        y -= 15
        for k, v in recettes.get('details', {}).items():
            c.drawString(120, y, f"{k.capitalize()}: {v}")
            y -= 15

        c.drawString(100, y, f"Dépenses totales: {depenses.get('total', 0)} FCFA")
        y -= 15
        for k, v in depenses.get('details', {}).items():
            c.drawString(120, y, f"{k.capitalize()}: {v}")
            y -= 15

        c.drawString(100, y, f"Budget prévisionnel: {budget.get('previsionnel', 0)} FCFA")
        y -= 15
        c.drawString(100, y, f"Budget réel: {budget.get('reel', 0)} FCFA")
        y -= 15
        c.drawString(100, y, f"Écart: {budget.get('ecart', 0)} FCFA")
        y -= 15
        c.drawString(100, y, f"Solde: {solde} FCFA")
        y -= 30

        # MATERIEL
        c.setFont("Helvetica-Bold", 14)
        c.drawString(100, y, "Rapport Matériel")
        y -= 20

        for section in ["materiels", "infrastructures"]:
            c.setFont("Helvetica-Bold", 12)
            c.drawString(100, y, section.capitalize())
            y -= 15
            total = data['materiel'][section].get("total", 0)
            c.setFont("Helvetica", 12)
            c.drawString(120, y, f"Total: {total}")
            y -= 15
            etats = data['materiel'][section].get("etat", {})
            c.drawString(120, y, "États:")
            y -= 15
            for etat, count in etats.items():
                c.drawString(140, y, f"{etat}: {count}")
                y -= 15
            y -= 20

        c.save()
        tmp_path = tmp.name

    return FileResponse(tmp_path, filename="rapport_audit_compile.pdf", media_type="application/pdf")






@router.get("/audit/compile/export-excel")
def export_excel_rapport_audit_compilé(
    date_debut: Optional[datetime] = None,
    date_fin: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, {"Administrateur", "Inspecteur"})
    if not date_debut or not date_fin:
        raise HTTPException(status_code=400, detail="date_debut et date_fin sont requis.")

    data = crud_rapport.generer_rapport_audit_compile(db, date_debut, date_fin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Audit Combiné"

    ws.append(["Période", f"{data['periode']['date_debut']} à {data['periode']['date_fin']}"])
    ws.append([])

    # Administratif
    ws.append(["Rapport Administratif", f"Total: {data['administratif']['total_rapports']}"])
    ws.append(["Titre", "Date Rapport", "Auteur", "Contenu"])
    for rapport in data['administratif']['rapports']:
        ws.append([rapport["titre"], rapport["date_rapport"], rapport["auteur"], rapport["contenu"]])

    ws.append([])

    # Financier
    ws.append(["Rapport Financier", f"Année: {data['periode']['annee_financiere']}"])
    ws.append(["Catégorie", "Montant (FCFA)"])
    recettes = data['financier'].get('recettes', {})
    depenses = data['financier'].get('depenses', {})
    budget = data['financier'].get('budget', {})
    solde = data['financier'].get('solde', 0)

    ws.append(["Recettes totales", recettes.get('total', 0)])
    for k, v in recettes.get('details', {}).items():
        ws.append([k.capitalize(), v])

    ws.append(["Dépenses totales", depenses.get('total', 0)])
    for k, v in depenses.get('details', {}).items():
        ws.append([k.capitalize(), v])

    ws.append(["Budget prévisionnel", budget.get('previsionnel', 0)])
    ws.append(["Budget réel", budget.get('reel', 0)])
    ws.append(["Écart", budget.get('ecart', 0)])
    ws.append(["Solde", solde])

    ws.append([])

    # Matériel
    for section in ["materiels", "infrastructures"]:
        ws.append([f"Rapport Matériel - {section.capitalize()}"])
        ws.append(["Total", data['materiel'][section].get("total", 0)])
        ws.append(["État", "Nombre"])
        for etat, count in data['materiel'][section].get("etat", {}).items():
            ws.append([etat, count])

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    return FileResponse(
        tmp_path,
        filename="rapport_audit_compile.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



@router.get("/materiel/export-pdf")
def export_pdf_materiel(
    date_debut: Optional[datetime] = None,
    date_fin: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, ALLOWED_ROLES_MATERIEL)
    rapport = crud_rapport.generer_rapport_materiel(db, date_debut=date_debut, date_fin=date_fin)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        c = canvas.Canvas(tmp.name)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(100, 800, "Rapport Matériels et Infrastructures")

        y = 780
        for section, data in rapport.items():
            c.setFont("Helvetica-Bold", 14)
            c.drawString(100, y, f"{section.capitalize()}:")
            y -= 20

            if "total" in data:
                c.setFont("Helvetica", 12)
                c.drawString(120, y, f"Total: {data['total']}")
                y -= 20

            if "etat" in data:
                c.setFont("Helvetica", 12)
                c.drawString(120, y, "États:")
                y -= 15
                for etat, count in data["etat"].items():
                    c.drawString(140, y, f"{etat}: {count}")
                    y -= 15

            if "pret" in data:
                c.setFont("Helvetica", 12)
                c.drawString(120, y, "Prêts en cours:")
                y -= 15
                for pret in data["pret"]:
                    pret_str = ", ".join(f"{k}: {v}" for k, v in pret.items())
                    c.drawString(140, y, pret_str)
                    y -= 15
                    if y < 100:
                        c.showPage()
                        y = 800
            y -= 20
            if y < 100:
                c.showPage()
                y = 800

        c.save()
        tmp_path = tmp.name

    return FileResponse(tmp_path, filename="rapport_materiels.pdf", media_type="application/pdf")


@router.get("/materiel/export-excel")
def export_excel_materiel(
    date_debut: Optional[datetime] = None,
    date_fin: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    check_role(current_user, ALLOWED_ROLES_MATERIEL)
    rapport = crud_rapport.generer_rapport_materiel(db, date_debut=date_debut, date_fin=date_fin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Matériels"

    # Matériels
    ws.append(["MATÉRIELS"])
    ws.append(["Total", rapport["materiels"].get("total", 0)])
    ws.append(["État", "Nombre"])
    for etat, count in rapport["materiels"].get("etat", {}).items():
        ws.append([etat, count])

    ws.append([])
    ws.append(["Prêts Matériels"])
    headers = ["Matériel", "Bénéficiaire", "Date prêt", "Date retour prévue", "Date retour effective", "État retour"]
    ws.append(headers)
    for pret in rapport["materiels"].get("pret", []):
        ws.append([
            pret.get("materiel", ""),
            pret.get("beneficiaire", ""),
            pret.get("date_pret", ""),
            pret.get("date_retour_prevue", ""),
            pret.get("date_retour_effective", ""),
            pret.get("etat_retour", "")
        ])

    ws.append([])
    # Infrastructures
    ws.append(["INFRASTRUCTURES"])
    ws.append(["Total", rapport["infrastructures"].get("total", 0)])
    ws.append(["État", "Nombre"])
    for etat, count in rapport["infrastructures"].get("etat", {}).items():
        ws.append([etat, count])

    ws.append([])
    ws.append(["Prêts Infrastructures"])
    headers = ["Infrastructure", "Bénéficiaire", "Date prêt", "Date retour prévue", "Date retour effective", "État retour"]
    ws.append(headers)
    for pret in rapport["infrastructures"].get("pret", []):
        ws.append([
            pret.get("infrastructure", ""),
            pret.get("beneficiaire", ""),
            pret.get("date_pret", ""),
            pret.get("date_retour_prevue", ""),
            pret.get("date_retour_effective", ""),
            pret.get("etat_retour", "")
        ])

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    return FileResponse(
        tmp_path,
        filename="rapport_materiels.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# --- RECHERCHE DE RAPPORTS ---

@router.get("/search", response_model=List[RapportOut])
def search_rapports(
    query: Optional[str] = None,
    type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    # Autorisation dynamique selon le type
    if type == "financier":
        check_role(current_user, ALLOWED_ROLES_FINANCIER)
    elif type == "administratif":
        check_role(current_user, ALLOWED_ROLES_ADMINISTRATIF)
    elif type == "materiel":
        check_role(current_user, ALLOWED_ROLES_MATERIEL)
    elif type == "audit":
        check_role(current_user, {"Administrateur", "Inspecteur"})
    else:
        # Si aucun type spécifié, l'utilisateur doit avoir au moins un rôle autorisé
        check_role(current_user, ALLOWED_ROLES_FINANCIER | ALLOWED_ROLES_ADMINISTRATIF | ALLOWED_ROLES_MATERIEL | {"Administrateur", "Inspecteur"})

    return crud_rapport.search_rapports(db, query=query, type=type)


